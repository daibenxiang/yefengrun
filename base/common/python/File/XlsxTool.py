import os, time, sys, os, json

from openpyxl import load_workbook, Workbook

sys.path.append("..")

from Log.LogController import LogController


class XlsxTool(object):
    
    def __init__(self, filename=None):
        self.class_name = "XlsxTool"
        self.ls = LogController.getLogger(self.class_name)
        if filename is None:
            self.filename = ""
            self.file_context = None
        else:
            self.filename = filename
            self.file_context = self.open()

    def __str__(self):
        return self.class_name

    def open(self, filename=None):
        self.ls.log_print("info", "open " + str(self.filename), "open")
        if filename is not None:
            self.filename = filename
        if not os.path.exists(self.filename):
            raise FileNotFoundError
        try:
            self.file_context = load_workbook(self.filename)
            return self.file_context
        except Exception as e:
            self.ls.log_print("error", str(e), "open")
            return False

    def getSheetNames(self):
        self.ls.log_print("info", "getSheetNames " + str(self.file_context), "getSheetNames")
        self.sheet_names = []
        try:
            self.sheet_names = self.file_context.sheetnames
            return self.sheet_names
        except Exception as e:
            self.ls.log_print("error", str(e), "getSheetNames")
            return False

    def readSheetbyName(self, sheet_name, *args):
        self.ls.log_print("info", "readSheetbyName " + str(sheet_name), "readSheetbyName")
        if not self.file_context is None:
            if sheet_name not in self.getSheetNames():
                return False
        else :
            return False
        
        sheet = None
        try:
            sheet = self.file_context[sheet_name]
            return sheet
        except Exception as e:
            self.ls.log_print("error", str(e), "readSheetbyName")
            return False
    def getRows(self, sheet):
        try:
            rows = sheet.rows
            return rows
        except Exception as e:
            self.ls.log_print("error", str(e), "getRowLength")
            return False

    def getColumns(self, sheet):
        try:
            rows = sheet.rows
            return rows
        except Exception as e:
            self.ls.log_print("error", str(e), "getRowLength")
            return False

    def getRowLength(self, sheet):
        try:
            columns = len(sheet.columns)
            return columns
        except Exception as e:
            self.ls.log_print("error", str(e), "getRowLength")
            return False

    def getColLength(self, sheet):
        try:
            col_length = len(sheet.columns)
            return col_length
        except Exception as e:
            self.ls.log_print("error", str(e), "getRowLength")
            return False

    def close(self):
        self.file_context.close()
        return False


if __name__ == "__main__":
    xt = XlsxTool("Template.xlsx")
    sheet_case = xt.readSheetbyName("testcases")
    print (sheet_case ["A1"].value)
    for x in sheet_case.rows:
        for x1 in x:
            print(x1.value)